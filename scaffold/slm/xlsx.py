# Copyright The Linux Foundation
# SPDX-License-Identifier: Apache-2.0

import os
from collections import OrderedDict
import openpyxl

from scaffold.datatypes import SLMLicense

##### Main xlsx reporting functions
##### External usage shouldn't require calling anything except these

def makeXlsx(categories):
    wb = openpyxl.Workbook()

    # look for "No license found" category, and if one exists, annotate it
    for cat in categories:
        if cat._name == "No license found":
            _annotateNoLicenseFound(cat)

    # create sheets
    _generateSummarySheet(wb, categories)
    _generateCategorySheets(wb, categories)
    _generateFileListings(wb, categories)

    # and return the workbook
    return wb

def saveXlsx(wb, path):
    try:
        wb.save(path)
        return True
    except PermissionError:
        return False

##### Helper functions for xlsx reporting

def _generateSummarySheet(wb, categories):
    # use the first (existing) sheet as the summary sheet
    ws = wb.active
    ws.title = "License summary"

    # adjust column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 10

    # create font styles
    fontBold = openpyxl.styles.Font(size=16, bold=True)
    fontNormal = openpyxl.styles.Font(size=14)
    alignNormal = openpyxl.styles.Alignment(wrap_text=True)

    # create headers
    ws['A1'] = "License"
    ws['A1'].font = fontBold
    ws['C1'] = "# of files"
    ws['C1'].font = fontBold

    # create category and license rows
    total = 0
    row = 3
    for cat in categories:
        if cat._numfiles <= 0:
            continue
        ws[f'A{row}'] = f'{cat._name}:'
        ws[f'A{row}'].font = fontBold
        row += 1
        for lic in cat._licenses:
            if lic._numfiles <= 0:
                continue
            ws[f'B{row}'] = lic._name
            ws[f'B{row}'].font = fontNormal
            ws[f'B{row}'].alignment = alignNormal
            ws[f'C{row}'] = lic._numfiles
            ws[f'C{row}'].font = fontNormal
            total += lic._numfiles
            row += 1

    # create total row
    row += 1
    ws[f'A{row}'] = "TOTAL"
    ws[f'A{row}'].font = fontBold
    ws[f'C{row}'] = total
    ws[f'C{row}'].font = fontBold

def _generateCategorySheets(wb, categories):
    # create font styles
    fontBold = openpyxl.styles.Font(size=16, bold=True)

    for cat in categories:
        # skip category if it has no files
        if cat._numfiles <= 0:
            continue

        ws = wb.create_sheet(cat._name)

        # and set column dimensions
        ws.column_dimensions['A'].width = 100
        ws.column_dimensions['B'].width = 60

        # and fill in sheet headers
        ws['A1'] = "File"
        ws['A1'].font = fontBold
        ws['B1'] = "License"
        ws['B1'].font = fontBold

def _generateFileListings(wb, categories):
    # create font styles
    fontNormal = openpyxl.styles.Font(size=14)
    alignNormal = openpyxl.styles.Alignment(wrap_text=True)

    for cat in categories:
        if cat._numfiles <= 0:
            continue
        ws = wb[cat._name]
        row = 2
        for lic in cat._licenses:
            if lic._numfiles <= 0:
                continue
            for fi in lic._files:
                ws[f'A{row}'] = fi._path
                ws[f'A{row}'].font = fontNormal
                ws[f'A{row}'].alignment = alignNormal
                ws[f'B{row}'] = lic._name
                ws[f'B{row}'].font = fontNormal
                ws[f'B{row}'].alignment = alignNormal
                row += 1

def _annotateNoLicenseFound(catNoLicense):
    licExt = None
    licEmpty = None
    licThird = None

    # for now, only looking at the first license in the category
    licNoLicense = catNoLicense._licenses[0]
    filesToRemove = []
    for fi in licNoLicense._files:
        labeled = False
        if not labeled:
            if fi._findings.get("thirdparty", "N/A") == "yes":
                # create new "license" if this is the first file we've seen for it
                if licThird is None:
                    licThird = SLMLicense()
                    licThird._name = "No license found - third party directory"
                    catNoLicense._licenses.append(licThird)
                # add file to this "license"
                licThird._files.append(fi)
                licThird._numfiles += 1
                labeled = True
                # and tag it to remove from "No license found"
                filesToRemove.append(fi)
                licNoLicense._numfiles -= 1
        if not labeled:
            if fi._findings.get("emptyfile", "N/A") == "yes":
                # create new "license" if this is the first file we've seen for it
                if licEmpty is None:
                    licEmpty = SLMLicense()
                    licEmpty._name = "No license found - empty file"
                    catNoLicense._licenses.append(licEmpty)
                # add file to this "license"
                licEmpty._files.append(fi)
                licEmpty._numfiles += 1
                labeled = True
                # and tag it to remove from "No license found"
                filesToRemove.append(fi)
                licNoLicense._numfiles -= 1
        if not labeled:
            if fi._findings.get("extension", "N/A") == "yes":
                # create new "license" if this is the first file we've seen for it
                if licExt is None:
                    licExt = SLMLicense()
                    licExt._name = "No license found - excluded file extension"
                    catNoLicense._licenses.append(licExt)
                # add file to this "license"
                licExt._files.append(fi)
                licExt._numfiles += 1
                labeled = True
                # and tag it to remove from "No license found"
                filesToRemove.append(fi)
                licNoLicense._numfiles -= 1

    # wait until we're done, so we can go back and remove them now
    licNoLicense._files = [x for x in licNoLicense._files if x not in filesToRemove]
