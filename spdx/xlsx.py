# Copyright The Linux Foundation
# SPDX-License-Identifier: Apache-2.0

import os
from collections import OrderedDict
import openpyxl


##### SPDX / Trivy xlsx reporting functions

def makeXlsx(spdxDocument):
    wb = openpyxl.Workbook()
    _generateDependenciesSheet(wb, spdxDocument)

    # and return the workbook
    return wb

def saveXlsx(wb, path):
    try:
        wb.save(path)
        return True
    except PermissionError:
        return False

##### Helper functions for xlsx reporting

def _generateDependenciesSheet(wb, spdxDocument):
    # use the first (existing) sheet as the dependencies sheet
    ws = wb.active
    ws.title = "Dependencies"

    # adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 50

    # create font styles
    fontBold = openpyxl.styles.Font(size=16, bold=True)
    fontNormal = openpyxl.styles.Font(size=14)
    alignNormal = openpyxl.styles.Alignment(wrap_text=True)

    # create headers
    ws['A1'] = "Package name"
    ws['A1'].font = fontBold
    ws['B1'] = "License"
    ws['B1'].font = fontBold
    ws['C1'] = "Version"
    ws['C1'].font = fontBold
    ws['D1'] = "Package URL"
    ws['D1'].font = fontBold
    ws['E1'] = "Dependency Relationship"
    ws['E1'].font = fontBold
    # Create dependncy rows
    # Collect all the package SPDX IDs
    unrecordedPackages = {}
    pkgRelationships = {}
    for pkg in spdxDocument.packages:
        unrecordedPackages[pkg.spdx_id] = pkg
        pkgRelationships[pkg.spdx_id] = [];
    docRelationships = []
    for relationship in spdxDocument.relationships:
        if relationship.spdx_element_id == spdxDocument.creation_info.spdx_id:
            docRelationships.append(relationship)
        elif relationship.spdx_element_id in unrecordedPackages:
            pkgRelationships[relationship.spdx_element_id].append(relationship)      
    for docRelationship in docRelationships:
        # start at the top so we sort in a somewhat natural hierarchy
        if docRelationship.related_spdx_element_id in unrecordedPackages:
            pkg = unrecordedPackages[docRelationship.related_spdx_element_id]
            _addrow(ws, pkg, "SPDXRef-DOCUMENT "+docRelationship.relationship_type.name)
            del unrecordedPackages[docRelationship.related_spdx_element_id]
            _addrelationships(ws, pkg, unrecordedPackages, pkgRelationships)
    for pkg in unrecordedPackages.values():
        _addrow(ws, pkg, "UNKNOWN - package in the document but not referenced in a relationship")
    
def _addrow(ws, pkg, relationship):
    name = pkg.name if pkg.name else ""
    version = pkg.version if pkg.version else ""
    if pkg.license_declared:
        declaredLicense = str(pkg.license_declared)
    else:
        declaredLicense = ""
    purl = ""
    for externalRef in pkg.external_references:
        if externalRef.reference_type == "purl":
            purl = externalRef.locator
    ws.append([name, declaredLicense, version, purl, relationship])

def _addrelationships(ws, pkg, unrecordedPackages, pkgRelationships):
     for relationship in pkgRelationships[pkg.spdx_id]:
        if relationship.related_spdx_element_id in unrecordedPackages:
            relatedPkg = unrecordedPackages[relationship.related_spdx_element_id]
            pkgName = pkg.name if pkg.name else "UNKNOWN"
            _addrow(ws, relatedPkg, pkgName + " " + relationship.relationship_type.name)
            del unrecordedPackages[relationship.related_spdx_element_id]
            _addrelationships(ws, relatedPkg, unrecordedPackages, pkgRelationships)
        